import openpyxl as xl
"""Make sure to include an excel file in the location of this file is"""
"""Input the same file name of the excel file"""

def get_workbook(excel_filename):
    """
    Handles loading an existing workbook or creating a new one if it doesn't exist.
    This function centralizes the file handling logic.
    """
    try:
        # Try to load the workbook
        wb = xl.load_workbook(f"{excel_filename}.xlsx")
        sheet = wb.active
        print(f"Loaded existing phone book: {excel_filename}.xlsx")
        wb.save(f"{excel_filename}.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new one with headers
        wb = xl.Workbook()
        sheet = wb.active
        # Correctly set headers in cells (1,1), (1,2), and (1,3)
        sheet.cell(1, 1, "Name")
        sheet.cell(1, 2, "Phone Number")
        sheet.cell(1, 3, "Email")
        print(f"File '{excel_filename}.xlsx' not found. A new one has been created.")
        wb.save(f"{excel_filename}.xlsx")

    return wb, sheet


def add_contact(sheet, excel_filename, wb):
    """
    Prompts the user for contact details and adds a new row to the sheet.
    """
    print("\n--- Add a New Contact ---")

    # Get the next available row number
    next_row = sheet.max_row + 1

    # Loop to add multiple contacts in one session
    while True:
        contact_name = input("Enter contact name: ")
        contact_number = input("Enter contact number: ")
        contact_email = input("Enter contact email: ")

        # Add the contact to the next row
        sheet.cell(next_row, 1, contact_name)
        sheet.cell(next_row, 2, contact_number)
        sheet.cell(next_row, 3, contact_email)

        # Increment the row counter for the next contact
        next_row += 1

        end_session = input("Do you want to add another contact? (Y/N): ").lower()
        if end_session == "n":
            break

    # Save the workbook after the loop finishes
    wb.save(f"{excel_filename}.xlsx")
    print("Contacts saved successfully.")


def search_contact(sheet):
    """
    Searches for a contact by name and prints all its details.
    """
    print("\n--- Search for a Contact ---")
    search_name = input("Enter the name you want to search for: ")
    found = False

    # Iterate through rows starting from the second row to skip headers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and search_name.lower() == str(row[0]).lower():
            # If a match is found, print all contact details from the row
            print(f"\nContact Found:")
            print(f"  Name: {row[0]}")
            print(f"  Phone Number: {row[1]}")
            print(f"  Email: {row[2]}")
            found = True
            break

    if not found:
        print(f"Contact '{search_name}' not found.")


def delete_contact(sheet, excel_filename, wb):
    """
    Searches for a contact by name and deletes the entire row after confirmation.
    """
    print("\n--- Delete a Contact ---")
    search_name = input("Enter the name you want to delete: ")
    row_to_delete = None

    # Iterate through rows by index to find the row number to delete
    for row_index in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=1).value
        # Check if the cell value is a case-insensitive match
        if cell_value and str(cell_value).lower() == search_name.lower():
            row_to_delete = row_index
            break

    if row_to_delete:
        verify = input(f"Are you sure you want to delete '{search_name}'? (Y/N): ").lower()
        if verify == "y":
            # Delete the entire row if confirmed
            sheet.delete_rows(row_to_delete)
            wb.save(f"{excel_filename}.xlsx")
            print(f"Contact '{search_name}' successfully deleted.")
        else:
            print("Deletion canceled.")
    else:
        print(f"Contact '{search_name}' not found. No contacts were deleted.")


def main_menu():
    """
    The main menu function to run the application.
    """
    excel_filename = input("Enter excel filename for your phone book: ")

    # Get or create the workbook and sheet
    wb, sheet = get_workbook(excel_filename)

    while True:
        print("\n--- Phone Book Menu ---")
        print("1. Add a new contact")
        print("2. Search for a contact")
        print("3. Delete a contact")
        print("4. Exit")

        choice = input("Enter your choice (1-4): ")

        if choice == "1":
            add_contact(sheet, excel_filename, wb)
        elif choice == "2":
            search_contact(sheet)
        elif choice == "3":
            delete_contact(sheet, excel_filename, wb)
        elif choice == "4":
            print("Exiting application. Goodbye!")
            break
        else:
            print("Invalid choice. Please enter a number from 1 to 4.")


if __name__ == "__main__":
    main_menu()
